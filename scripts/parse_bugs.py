"""缺陷数据统计脚本（通用工具，所有迭代版本通用）

用法：
    python scripts/parse_bugs.py <文件路径>

    # 直接读 Excel（推荐）
    python scripts/parse_bugs.py 630版本/iFLYBUDS--2026-04-16.xlsx

    # 或读预先转好的 JSON
    python scripts/parse_bugs.py 630版本/excel_data2.json

输入：
    - .xlsx：缺陷导出 Excel，需包含名为 "work items" 的 sheet，第一行为表头
    - .json：根节点需包含 "work items" 数组，第一行为表头，后续每行一条记录

输出：
    标准输出打印四张统计表 —— 总缺陷数、各端数量、严重程度×状态、各端×优先级、
    暂不修复问题列表，可直接填充到测试报告对应章节。

依赖：
    openpyxl（仅读 .xlsx 时需要）  pip install openpyxl
"""

import json
import os
import re
import sys
from collections import defaultdict


def get_platform(title):
    match = re.match(r'[【\[]([^】\]]+)[】\]]', title)
    if not match:
        return '未知'
    tag = match.group(1).upper()
    if '服务端' in tag or 'SERVER' in tag:
        return '服务端'
    if 'ANDROID' in tag:
        return 'Android'
    if 'IOS' in tag:
        return 'iOS'
    return '未知'


def get_severity(s):
    s = (s or '').strip()
    if '致命' in s or s.startswith('1'):
        return '致命'
    elif '严重' in s or s.startswith('2'):
        return '严重'
    elif '一般' in s or s.startswith('3'):
        return '一般'
    elif '轻微' in s or s.startswith('4'):
        return '轻微'
    return s or '未知'


def get_status(s):
    s = (s or '').strip()
    if s in ('已关闭', '已关闭（未修复）'):
        return '已闭环'
    elif s == '暂不修复':
        return '暂不修复'
    elif s in ('待修复', '处理中', '重新打开'):
        return '待修复'
    return s


def load_rows_from_json(path):
    with open(path, encoding='utf-8') as f:
        data = json.load(f)
    return data['work items']


def load_rows_from_xlsx(path):
    try:
        import openpyxl
    except ImportError:
        print('错误：读取 .xlsx 需要 openpyxl，执行 `pip install openpyxl` 后重试', file=sys.stderr)
        sys.exit(1)
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_name = 'work items' if 'work items' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(['' if c is None else str(c) for c in row])
    return rows


def load_bugs(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == '.json':
        rows = load_rows_from_json(path)
    elif ext in ('.xlsx', '.xlsm'):
        rows = load_rows_from_xlsx(path)
    else:
        print(f'错误：不支持的文件类型 {ext}，仅支持 .xlsx / .xlsm / .json', file=sys.stderr)
        sys.exit(1)

    if not rows:
        return []
    headers = rows[0]
    bugs = []
    for row in rows[1:]:
        d = dict(zip(headers, row))
        if d.get('工作项类型', '') not in ('缺陷', '缺陷（产品/设计）'):
            continue
        bugs.append({
            'title':    d.get('标题', ''),
            'status':   get_status(d.get('状态', '')),
            'priority': (d.get('优先级', '') or '').strip(),
            'severity': get_severity(d.get('严重程度', '')),
            'platform': get_platform(d.get('标题', '')),
        })
    return bugs


def print_report(bugs):
    total = len(bugs)
    print(f'总缺陷数: {total}')
    print()

    platform_count = defaultdict(int)
    for b in bugs:
        platform_count[b['platform']] += 1
    print('=== 各端数量 ===')
    for p in ['Android', 'iOS', '服务端', '未知']:
        if platform_count[p]:
            print(f'  {p}: {platform_count[p]}')
    print()

    print('=== 严重程度 × 状态 ===')
    sev_stat = defaultdict(lambda: defaultdict(int))
    for b in bugs:
        sev_stat[b['severity']][b['status']] += 1

    statuses   = ['暂不修复', '待修复', '已闭环']
    severities = ['致命', '严重', '一般', '轻微']
    print('严重程度\t暂不修复\t待修复\t已闭环\t合计')
    for sev in severities:
        vals = [sev_stat[sev][st] for st in statuses]
        print(f'{sev}\t' + '\t'.join(str(v) for v in vals) + f'\t{sum(vals)}')
    col_totals = [sum(sev_stat[sev][st] for sev in severities) for st in statuses]
    print(f'合计\t' + '\t'.join(str(v) for v in col_totals) + f'\t{sum(col_totals)}')
    print()

    print('=== 各端 × 优先级 ===')
    plat_pri = defaultdict(lambda: defaultdict(int))
    for b in bugs:
        plat_pri[b['platform']][b['priority']] += 1

    priorities = ['紧急', '高', '中', '低']
    platforms  = ['Android', 'iOS', '服务端']
    print('端\t紧急\t高\t中\t低\t合计\t占比')
    for plat in platforms:
        vals = [plat_pri[plat][p] for p in priorities]
        t = sum(vals)
        pct = f'{t / total * 100:.1f}%' if total else '0.0%'
        print(f'{plat}\t' + '\t'.join(str(v) for v in vals) + f'\t{t}\t{pct}')
    print()

    print('=== 暂不修复问题列表 ===')
    not_fix = [b for b in bugs if b['status'] == '暂不修复']
    print(f'共 {len(not_fix)} 个')
    for i, b in enumerate(not_fix, 1):
        print(f'{i}. [{b["platform"]}][{b["severity"]}][{b["priority"]}] {b["title"]}')


def main():
    if len(sys.argv) < 2:
        print('用法：python scripts/parse_bugs.py <文件路径>', file=sys.stderr)
        print('  支持 .xlsx / .xlsm / .json', file=sys.stderr)
        sys.exit(1)

    path = sys.argv[1]
    if not os.path.isfile(path):
        print(f'错误：找不到文件 {path}', file=sys.stderr)
        sys.exit(1)

    bugs = load_bugs(path)
    print_report(bugs)


if __name__ == '__main__':
    main()
